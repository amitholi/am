"""Bulk importer for CSV and Excel files"""
import csv
import asyncio
from typing import Optional
import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import AsyncSessionLocal
from app.models.case import Case
from app.models.court import Court
from app.models.party import Party

logger = structlog.get_logger()

DEFAULT_MAPPING = {
    "case_number": ["מספר תיק", "case_number", "תיק"],
    "case_type": ["סוג הליך", "case_type", "סוג"],
    "court": ["בית משפט", "court", "court_name"],
    "filing_date": ["תאריך הגשה", "filing_date", "תאריך"],
    "status": ["סטטוס", "status"],
    "judge_name": ["שופט", "judge", "judge_name"],
    "description": ["תיאור", "description", "נושא"],
    "plaintiff": ["תובע", "מבקש", "plaintiff"],
    "defendant": ["נתבע", "משיב", "defendant"],
}


def _find_field(row: dict, candidates: list) -> Optional[str]:
    """Find a field value by checking multiple candidate column names"""
    for key in candidates:
        if key in row and row[key] is not None and str(row[key]).strip():
            return str(row[key]).strip()
    return None


def _parse_date(value: Optional[str]):
    if not value:
        return None
    from datetime import date
    import re
    value = str(value).strip()
    try:
        from datetime import datetime
        return datetime.fromisoformat(value[:10]).date()
    except Exception:
        pass
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", value)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except Exception:
            pass
    return None


class BulkImporter:
    """Bulk importer for CSV and Excel files"""

    async def import_csv(self, file_path: str, mapping: dict = None) -> int:
        """Import cases from CSV file"""
        if mapping is None:
            mapping = DEFAULT_MAPPING

        logger.info("csv_import_start", file=file_path)
        total = 0

        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        async with AsyncSessionLocal() as db:
            for row in rows:
                added = await self._save_row(db, row, mapping)
                if added:
                    total += 1
                if total % 100 == 0:
                    await db.commit()
                    logger.info("csv_import_progress", total=total)
            await db.commit()

        logger.info("csv_import_done", total=total)
        return total

    async def import_excel(self, file_path: str, sheet: str = None, mapping: dict = None) -> int:
        """Import cases from Excel file"""
        import openpyxl
        if mapping is None:
            mapping = DEFAULT_MAPPING

        logger.info("excel_import_start", file=file_path)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return 0

        headers = [str(h).strip() if h else "" for h in rows[0]]
        total = 0

        async with AsyncSessionLocal() as db:
            for row_values in rows[1:]:
                row = {headers[i]: str(v).strip() if v is not None else "" for i, v in enumerate(row_values)}
                added = await self._save_row(db, row, mapping)
                if added:
                    total += 1
                if total % 100 == 0:
                    await db.commit()
            await db.commit()

        logger.info("excel_import_done", total=total)
        return total

    async def _save_row(self, db: AsyncSession, row: dict, mapping: dict) -> bool:
        case_number = _find_field(row, mapping.get("case_number", []))
        if not case_number:
            return False

        court_name = _find_field(row, mapping.get("court", []))
        court_id = None
        if court_name:
            result = await db.execute(
                select(Court).where(Court.name.ilike(f"%{court_name[:15]}%")).limit(1)
            )
            court = result.scalar_one_or_none()
            if court:
                court_id = court.id

        existing = await db.execute(
            select(Case).where(Case.case_number == case_number, Case.court_id == court_id)
        )
        if existing.scalar_one_or_none():
            return False

        case_type_raw = _find_field(row, mapping.get("case_type", []))

        case = Case(
            case_number=case_number,
            case_type=case_type_raw or "אחר",
            court_id=court_id,
            filing_date=_parse_date(_find_field(row, mapping.get("filing_date", []))),
            status=_find_field(row, mapping.get("status", [])) or "לא ידוע",
            judge_name=_find_field(row, mapping.get("judge_name", [])),
            description=_find_field(row, mapping.get("description", [])),
            raw_data=dict(row),
        )
        db.add(case)
        await db.flush()

        plaintiff = _find_field(row, mapping.get("plaintiff", []))
        if plaintiff:
            db.add(Party(case_id=case.id, name=plaintiff[:500], party_type="תובע"))

        defendant = _find_field(row, mapping.get("defendant", []))
        if defendant:
            db.add(Party(case_id=case.id, name=defendant[:500], party_type="נתבע"))

        return True
