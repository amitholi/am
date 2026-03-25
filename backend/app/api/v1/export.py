from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional
import csv
import io
import json

from app.database import get_db
from app.models.case import Case
from app.models.court import Court

router = APIRouter()


async def _fetch_cases(db: AsyncSession, filters: dict, limit: int = 10000):
    query = (
        select(Case, Court.name.label("court_name"))
        .outerjoin(Court, Case.court_id == Court.id)
    )
    conditions = []
    from sqlalchemy import and_
    if filters.get("case_type"):
        conditions.append(Case.case_type == filters["case_type"])
    if filters.get("court_id"):
        conditions.append(Case.court_id == int(filters["court_id"]))
    if filters.get("status"):
        conditions.append(Case.status == filters["status"])
    if filters.get("date_from"):
        conditions.append(Case.filing_date >= filters["date_from"])
    if filters.get("date_to"):
        conditions.append(Case.filing_date <= filters["date_to"])
    if conditions:
        query = query.where(and_(*conditions))
    query = query.order_by(Case.filing_date.desc().nullslast()).limit(limit)
    result = await db.execute(query)
    return result.all()


@router.get("/export/csv")
async def export_csv(
    case_type: Optional[str] = Query(None),
    court_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_cases(db, {
        "case_type": case_type, "court_id": court_id,
        "status": status, "date_from": date_from, "date_to": date_to
    })

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["מספר תיק", "סוג הליך", "בית משפט", "תאריך הגשה", "סטטוס", "שופט", "תיאור"])
    for row in rows:
        case = row[0]
        court_name = row[1] or ""
        writer.writerow([
            case.case_number, case.case_type, court_name,
            str(case.filing_date) if case.filing_date else "",
            case.status or "", case.judge_name or "", case.description or "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=cases.csv"},
    )


@router.get("/export/json")
async def export_json(
    case_type: Optional[str] = Query(None),
    court_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    rows = await _fetch_cases(db, {
        "case_type": case_type, "court_id": court_id,
        "status": status, "date_from": date_from, "date_to": date_to
    })

    data = []
    for row in rows:
        case = row[0]
        data.append({
            "id": case.id,
            "case_number": case.case_number,
            "case_type": case.case_type,
            "case_type_description": case.case_type_description,
            "court_name": row[1],
            "filing_date": str(case.filing_date) if case.filing_date else None,
            "status": case.status,
            "judge_name": case.judge_name,
            "description": case.description,
        })

    json_str = json.dumps(data, ensure_ascii=False, indent=2, default=str)
    return StreamingResponse(
        iter([json_str]),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=cases.json"},
    )


@router.get("/export/excel")
async def export_excel(
    case_type: Optional[str] = Query(None),
    court_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    rows = await _fetch_cases(db, {
        "case_type": case_type, "court_id": court_id,
        "status": status, "date_from": date_from, "date_to": date_to
    })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "תיקים"
    ws.sheet_view.rightToLeft = True

    headers = ["מספר תיק", "סוג הליך", "תיאור הליך", "בית משפט", "תאריך הגשה", "סטטוס", "שופט", "תיאור"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")

    for i, row in enumerate(rows, 2):
        case = row[0]
        ws.cell(row=i, column=1, value=case.case_number)
        ws.cell(row=i, column=2, value=case.case_type)
        ws.cell(row=i, column=3, value=case.case_type_description)
        ws.cell(row=i, column=4, value=row[1])
        ws.cell(row=i, column=5, value=str(case.filing_date) if case.filing_date else "")
        ws.cell(row=i, column=6, value=case.status)
        ws.cell(row=i, column=7, value=case.judge_name)
        ws.cell(row=i, column=8, value=case.description)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cases.xlsx"},
    )
